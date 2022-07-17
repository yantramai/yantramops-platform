import ast

import numpy as np
import pandas as pd
import traceback
import pathlib, sys, os, requests, json
from openpyxl import load_workbook
import os.path
from os import path
import random, string
from datetime import datetime, timedelta

_ROOT = os.path.abspath(os.path.dirname(__file__))
DATE_FORMAT = "%Y-%m-%d %H:%M:%S"


class Utils:
    def __init__(self, source_file):
        # Initialise parser for command-line arguments
        self.source_file = source_file
        extension = pathlib.Path(source_file).suffix
        stem = pathlib.Path(source_file).stem
        self.file_ext = extension
        self.stem = stem
        self.df = None
        self.encoding = 'latin1'

    def write_to_file(self, df, custom_file_extension=None, prefix_type=None, sheet_name='Default Sheet',
                      custom_dir='default', custom_output_file_name='default_file_name', source_file=""):
        # Writing the output to the current location
        if source_file == '.DS_Store':
            return
        stem = custom_output_file_name
        if not custom_output_file_name:
            stem = pathlib.Path(source_file).stem

        parent = os.path.join(pathlib.Path(source_file).parent, custom_dir)
        if not path.exists(parent):
            os.mkdir(parent)

        if custom_file_extension:
            custom_output_file = os.path.join(parent, stem + custom_file_extension)
            self.output_file = custom_output_file
        else:
            custom_file_extension = self.file_ext
            custom_output_file = os.path.join(parent, stem + custom_file_extension)
            self.output_file = custom_output_file

        try:
            if custom_file_extension == '.csv':
                df.to_csv(path_or_buf=custom_output_file, sep=',', encoding=self.encoding, index=False, header=True)
            if custom_file_extension == '.json':
                # df.to_json(path_or_buf=custom_output_file, orient="records", date_format="yyyy-MM-dd HH:mm:ss")
                df.to_json(path_or_buf=custom_output_file, orient="records")
            if custom_file_extension == '.xlsx':
                self.append_df_to_excel(df, sheet_name=sheet_name)

        except Exception as e:
            sys.stdout.write("Error writing data to file!\n\n" + str(e) + "\n")
            sys.exit(1)
        return self.output_file

    def append_df_to_excel(self, df, sheet_name='Default Sheet', startrow=None, startcol=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param output_file: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('d:/temp/test.xlsx', df)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False, startrow=25)

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        # Excel file doesn't exist - saving and exiting

        if not os.path.isfile(self.output_file):
            df.to_excel(
                self.output_file,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                startcol=startcol if startcol is not None else 0,

                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(self.output_file, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(self.output_file)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

            # write out the new sheet
            try:
                df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
            except ValueError:
                print("Oops!  there was an error in writing to the sheet.  Try again..." + ValueError)
            except Exception as e:
                print("Oops!  there was an error in writing to the sheet.  Try again..." + e)

        # save the workbook
        writer.save()

    def retrieve_dataframe(self, file_name, sheet_name, encoding):
        file_extention = pathlib.Path(file_name).suffix
        # Loading the input file
        try:
            if file_extention == '.json':
                self.df = pd.read_json(file_name, encoding=encoding, dtype='str',lines=True)
            if file_extention == '.csv':
                self.df = pd.read_csv(file_name, encoding=encoding, dtype='str')
            if file_extention == '.xlsx' or file_extention == '.xls':
                try:
                    self.df = pd.read_excel(file_name,sheet_name=sheet_name)
                except Exception as e:
                    traceback.print_exc()
                    sys.stdout.write(str(e) + "\n")
                    sys.exit(1)
            # sys.stdout.write("Successfully read the file: " + file_name + "\n")
        except FileNotFoundError as f:
            sys.stdout.write(str(f) + "\n")
            sys.exit(1)
        return self.df

    def get_config_data(self, data):
        software_scanning_file = data['software_scanning_file']
        file_name = data['file_name']
        encodings = data['encoding']
        drive_path_input = data['drive_path_input']
        drive_path_output = data['drive_path_output']
        excel_sheet_name = data['excel_sheet_name']
        identifier = data['columns_to_process']['identifier']
        title = data['columns_to_process']['title']
        description = data['columns_to_process']['description']
        write_me = data['write_to_excel']
        display_columns = data['display_columns']
        self.encoding = encodings
        return title, description, identifier, drive_path_input, drive_path_output, excel_sheet_name, file_name, write_me, software_scanning_file, display_columns

    def get_random_id(self):
        return 'vz'.join(random.choices(string.ascii_letters + string.digits, k=20))

    def get_data(self, path):
        return os.path.join(_ROOT, 'data', path)

    def print_columns_in_datasource(self, datasource):
        columns = datasource.columns.unique()
        for column in columns:
            print(column)

    def print_unique_values_in_columns(self, datasource, column_name):
        columns = datasource[column_name].unique().tolist()
        return columns

    def print_missing_columns(self, source_datasource, destination_datasource):
        source_datasource_column_list = source_datasource.columns.to_list()
        destination_datasource_column_list = destination_datasource.columns.to_list()
        list_difference = [column for column in source_datasource_column_list if
                           column not in destination_datasource_column_list]
        print(list_difference)

    def print_count_of_columns(self, source_datasource, column_names):
        print(source_datasource[column_names].isnull().sum())

    def print_empty_values(self, source_datasource, column_name):
        false_ = source_datasource[source_datasource[column_name].apply(lambda x: x.isspace() == False)]
        return false_

    def return_rows_matching_values(self, source_datasource, column_name, value):
        return source_datasource[source_datasource[column_name].apply(lambda x: str(x).find(value) != -1)]

    def return_rows_matching_values_in_list(self, source_datasource, column_name, values):
        return source_datasource[source_datasource[column_name].apply(lambda x: x in values)]

    def return_rows_not_matching_value(self, source_datasource, column_name, value):
        return source_datasource[source_datasource[column_name].apply(lambda x: str(x).find(value) == -1)]

    def return_rows_between_dates(self, source_datasource, column_name, start_date, end_date):
        return source_datasource[source_datasource[column_name].apply(lambda x: end_date >= x >= start_date)]

    def print_sorted_record_cound(self, source_datasource, column_name):
        print(source_datasource.count().sort_values())

    def cleanup_dates(self):
        xlsx = '/Users/jayantkaushal/Data/Repo/POCUtils/data/output/VERIZON_NP_SOE_AUDIT_LOGS_AUDIT_20211029_165310.xlsx'
        audit_log_data_frame = self.utils1.retrieve_dataframe(file_name=xlsx, sheet_name='Default Sheet',
                                                              encoding='latin1')
        audit_log_data_frame['_source_response'] = audit_log_data_frame['_source_response'].apply(
            lambda x: str(x).strip())
        audit_log_data_frame.update(
            audit_log_data_frame.select_dtypes('datetime').apply(lambda x: x.dt.strftime("%Y-%m-%d %H:%M:%S")))
        # audit_log_data_frame['_source_log_time_stamp'] = audit_log_data_frame['_source_log_time_stamp'].apply(lambda x: str(x).strip())
        audit_log_data_frame = audit_log_data_frame[audit_log_data_frame['_source_log_time_stamp'] != ' ']
        audit_log_data_frame['_source_log_time_stamp'] = audit_log_data_frame['_source_log_time_stamp'].apply(
            lambda x: pd.to_datetime(x).strftime('%Y-%m-%d %H:%M:%S'))
        audit_log_data_frame['_source_@timestamp'] = audit_log_data_frame['_source_log_time_stamp'].apply(
            lambda x: pd.to_datetime(x).strftime('%Y-%m-%d %H:%M:%S'))
        output_file = self.utils1.write_to_file(audit_log_data_frame, custom_file_extension='.xlsx')
        print(output_file)

    def invoke_REST_call(self):
        url = "https://auth.pingone.com/512a7bbf-abae-473f-8b71-06adbc1b3433/as/authorize?response_type=token&response_mode=pi.flow&scope=profile&redirect_uri=http://localhost&login_hint_token=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJjM2JhOGYwNi0zNGZlLTRkZGItOTZjNC1jNDE0MzhjNzQzOTYiLCJhdWQiOiJodHRwczovL2F1dGgucGluZ29uZS5jb20vNTEyYTdiYmYtYWJhZS00NzNmLThiNzEtMDZhZGJjMWIzNDMzL2FzIiwic3ViIjoiOGI0MTliZTAtOWVjMy00YTEwLTk2NjAtODVkMzE2YmE5NjAxIiwiaWF0IjoxNjMxMDE3NjIxLCJleHAiOjE2MzExMDQzMjF9.Ahqg1V5-952ryjXzNrelXNk13EgPCV_hU8ntKzyO9r0"
        url1 = "curl --location --request GET 'https://auth.pingone.com/512a7bbf-abae-473f-8b71-06adbc1b3433/as/authorize?response_type=token&response_mode=pi.flow&scope=profile&redirect_uri=http://localhost&login_hint_token=eyJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJjM2JhOGYwNi0zNGZlLTRkZGItOTZjNC1jNDE0MzhjNzQzOTYiLCJzdWIiOiI3ODkwMDhiMC1hNTE1LTRlOTctYmU0OS03ZjE2MmRjYTFhOWIiLCJpYXQiOjE2MzA5Mjk3ODMsImV4cCI6MTYzMDkzNjk4MywiYXVkIjoiaHR0cHM6Ly9hdXRoLnBpbmdvbmUuY29tLzUxMmE3YmJmLWFiYWUtNDczZi04YjcxLTA2YWRiYzFiMzQzMy9hcyJ9.j5w3p"
        payload = {}
        headers = {}
        response = requests.request("GET", url, headers=headers, data=payload)

    verizon_jira_ticket_columns = [
        'Summary',
        'Components',
        'Root Cause',
        'Application',
        'EnvironmentDetected',
        'ApplicationList',
        'Domain',
        'Status'
    ]
    focused_new_relic_columns = [
        '_messagetimems',
        '_messagetime',
        'type',
        '_collector',
        'computer',
        'instace',
        '_source',
        '_sourcehost',
        'title',
        '_raw',
    ]
    focused_new_relic_columns_1 = [
        'title'
        # '_raw',
    ]

    def get_bj_title_from_raw_message(self, raw_message, result_message):
        try:
            if result_message != None:
                return result_message
            a_dictionary = json.loads(raw_message)
            message = str(a_dictionary['message'])
            return message
        except:
            return raw_message

    def get_bj_date_time_from_ms(self, message_time):
        # print('12/17/2021 15:39:17.479 +0530')
        # message_time1 = datetime.strptime(str(message_time), '%Y/%m/%d %H:%M:%S.%f %z')
        message_time1 = datetime.strptime(str(message_time), '%m/%d/%Y %H:%M:%S.%f %z')
        past_date_before_2hours = message_time1 + timedelta(hours=5, minutes=30)
        extracted_date = past_date_before_2hours.strftime('%m/%d/%Y %H:%M:%S')
        # message_time3 = datetime.strptime(str(past_date_before_2hours), '%m/%d/%Y %H:%M:%S.%f%z')
        # print(message_time3)
        return extracted_date

    def container_name(self, kubernetes1):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        return kubernetes['container_name']

    def retrieve_container_key(self, kubernetes1, key):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        if key in kubernetes:
            return kubernetes[key]

    def namespace_name(self, kubernetes1):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        return kubernetes['namespace_name']

    def pod_name(self, kubernetes1):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        return kubernetes['pod_name']

    def container_image_id(self, kubernetes1):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        return kubernetes['container_image_id']

    def pod_id(self, kubernetes1):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        return kubernetes['pod_id']

    def host(self, kubernetes1):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        return kubernetes['host']

    def namespace_id(self, kubernetes1):
        tesval = json.dumps(ast.literal_eval(kubernetes1))
        kubernetes = json.loads(tesval)
        return kubernetes['namespace_id']

    def convert_et_to_GMT(self, date_and_time):
        time_change = timedelta(hours=5)
        new_time = date_and_time + time_change
        return new_time

    def get_random_id(self):
        return ''.join(random.choices(string.ascii_letters + string.digits, k=20))


    def get_matching_alert_with_id(self, id,matching_id,row,file):
        if id == matching_id:
            print('****')
            print(file)
            print(row)
            print('****')

            # print(file)
            # print(row)

    def get_audit_log_title(self, status_code, response, requestURI):
        if (str(status_code).strip() != '200') and len(str(response).strip()) != 0:

            try:
                a_dictionary = json.loads(response)
                if "error" in a_dictionary:
                    if type(a_dictionary['error']) == list and 'errorSpec' in a_dictionary['error'][0]:
                        error_ = a_dictionary['error'][0]['errorSpec']['name']
                        return error_
                    if 'MAX_DEVICE_LIMIT_EXCEEDED_ERROR' in a_dictionary['error']:
                        error_ = a_dictionary['error']['MAX_DEVICE_LIMIT_EXCEEDED_ERROR']
                        return error_

                if "errors" in a_dictionary:
                    if "errorMsg" in a_dictionary['errors']:
                        return (a_dictionary['errors']['errorMsg'])
                    if "message" in a_dictionary['errors'][0]:
                        message_ = a_dictionary['errors'][0]['message']
                        return message_
                    else:
                        return a_dictionary['errors'][0]['name'] + '(' + a_dictionary['errors'][0][
                            'debug_id'] + ')' + '\tnamespace:' + \
                               a_dictionary['errors'][0]['namespace']

                if "fault" in a_dictionary:
                    return a_dictionary['fault']['faultstring'] + '[' + a_dictionary['fault']['detail'][
                        'errorcode'] + ']'
                if "errorMsg" in a_dictionary:
                    if "RequestError" in a_dictionary['errorMsg']:
                        import re
                        myString = str(a_dictionary['errorMsg'])
                        myString = re.sub(r"[\n\t\s]*", "", myString).strip()
                        res = json.loads(myString)
                        return ('ExceptionType : ' + res['RequestError']['ExceptionType'] + '\t' + 'MessageId:' +
                                res['RequestError']['MessageId'] + '\t' + 'Text:' + res['RequestError']['Text'])
                    else:
                        return (a_dictionary['errorMsg'])

                if "serviceHeader" in a_dictionary:
                    return a_dictionary['serviceHeader']['statusMsg']

                if "data" in a_dictionary:
                    return requestURI
                if "meta" in a_dictionary:
                    return a_dictionary['meta']['cxpCorrelationId']
                if "serviceBody" in a_dictionary:
                    if 'cartValidationErrors' in a_dictionary['serviceBody']['serviceResponse']['context'][
                        'cartInfo']:
                        validation_errors = a_dictionary['serviceBody']['serviceResponse']['context']['cartInfo'][
                            'cartValidationErrors']
                        error_messages = ''
                        if (len(validation_errors) != 0):
                            for validation_error in validation_errors:
                                error_messages += validation_error['errorMessage'] + '\n'
                        return error_messages
                    else:
                        return a_dictionary['serviceBody']['serviceResponse']['context']['caseId']
            except:
                if response == 'None':
                    return requestURI
                return response
        else:
            uri_ = len(str(requestURI).strip()) == 0
            if uri_:
                return response

        return requestURI

    def get_matching_alert_with_content(self, content,row):
        if content == "Created Date Time":
            print('****')
            print(row)
            print('****')

            # print(file)
            # print(row)
